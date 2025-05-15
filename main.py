# Стандартные библиотеки
import os
import random
import secrets
import string
import sys
import time
import traceback
from datetime import datetime

# Внешние библиотеки
import asyncio
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

from automation.run_automation import schedule_next_run, check_auto_mode
from database import process_activity, DatabaseError, process_random_profile

# Локальные модули
from lava_moat import modify_file_runtimelavamoat
from meta_mask import MetaMaskHelper
from create_mm_wallet import create_wallet
from config import (
    logger, DATA_BASE_PATH, WORKSHEET_NAME,
    MODE_CLOSE_PROFILE, GLOBAL_SETTINGS, MIX_PROFILES, PROFILE_DELAY, AUTO_MODE
)
from MoreLogin.browser_manager import BrowserManager

# Проверяем режим работы в начале выполнения
check_auto_mode()

def create_password():
    """Создает надежный пароль длиной 22 символа, включающий буквы, цифры и специальные символы."""
    password = f"{''.join(secrets.choice(string.ascii_letters + string.digits + string.punctuation) for i in range(22))}"
    return password


class MainError(Exception):
    """Базовый класс для ошибок основного скрипта"""
    pass


def workbook_worksheet():
    """Инициализация и получение рабочей книги и листа Excel."""
    try:
        workbook_mm = openpyxl.load_workbook(DATA_BASE_PATH)
        worksheet_mm = workbook_mm[WORKSHEET_NAME]
        logger.debug(
            f" (workbook_worksheet) DATABASE Рабочий лист базы данных: {worksheet_mm.title}"
        )
        return workbook_mm, worksheet_mm
    except KeyError:
        logger.error(
            f" (workbook_worksheet) Error: Нет доступа к рабочему листу базы данных: {WORKSHEET_NAME}\n"
            f"Проверьте файл переменных окружения .env:\n"
            f" WORKSHEET_NAME \n"
        )
        raise MainError(f"Worksheet {WORKSHEET_NAME} not found")
    except InvalidFileException:
        logger.error(
            f" (workbook_worksheet) Error: Неверный формат файла базы данных: {DATA_BASE_PATH}\n"
            f"Проверьте файл переменных окружения .env:\n"
            f" DATA_BASE \n"
        )
        raise MainError(f"Invalid database file format: {DATA_BASE_PATH}")


async def restart_browser_profile(driver, env_id, unique_id, env_name, count):
    """Перезапуск профиля браузера при ошибках"""
    if count > 3:
        logger.critical(
            f"\n{'#' * 20} (main_flow), (operationEnv), Не удачный запуск №: {count}! Profile №: {unique_id}, Env_Name: {env_name}, Env ID: {env_id},"
            f"Env ID: {env_id} будет остановлен и закрыт. {'#' * 20}"
        )
        raise MainError(f"Failed to start browser profile after {count} attempts")
    else:
        # Закрываем драйвер и останавливаем профиль
        await asyncio.sleep(5)
        if driver:
            try:
                driver.quit()
            except Exception as e:
                logger.warning(f"Error closing driver: {e}")
        await BrowserManager.stop_browser_profile(env_id)
        logger.warning(
            f"\n (main_flow), (operationEnv) Не удачный запуск №: {count}! Повторный запуск через 5 секунд! Profile №: {unique_id}, Env_Name: {env_name}, Env ID: {env_id}"
        )
    return count


async def main_flow(
    env_id,
    seed,
    password,
    env_name,
    unique_id,
    mm_address,
    worksheet_mm,
    workbook_mm,
    row,
    mode_close_profile_or_not="n"
):
    """Основной рабочий процесс для одного профиля."""
    driver = None
    count = 0
    try:
        while True:
            if count <= 3:
                await asyncio.sleep(5)
            # Запуск профиля
            debug_url, driver_path = await BrowserManager.start_browser_profile(env_id)
            if not debug_url and not driver_path:
                count += 1
                count = await restart_browser_profile(driver, env_id, unique_id, env_name, count)
                continue

            # Логирование запуска профиля
            logger.warning(
                f"\n{'#' * 20} (main_flow) SCRIPT STARTED Profile №: {unique_id}, Env_Name: {env_name}, Env ID: {env_id} {'#' * 20}\n"
            )

            # Создание драйвера
            try:
                driver = await BrowserManager.create_web_driver(debug_url, driver_path)
                # Устанавливаем глобальное время ожидания в 10 секунды
                driver.implicitly_wait(10)
            except Exception as e:
                logger.error(f"Error creating web driver: {e}")
                count += 1
                count = await restart_browser_profile(driver, env_id, unique_id, env_name, count)
                continue

            # Основные операции
            logger.debug(
                f" (main_flow) STEP 1 <<< Начало работы в Profile №: {unique_id}, Env_Name: {env_name}, Env ID: {env_id} >>>"
            )

            try:
                await operationEnv(
                    driver, seed, env_id, password, mm_address,
                    worksheet_mm, workbook_mm, row, DATA_BASE_PATH
                )
                break
            except Exception as e:
                logger.error(f"Error in operationEnv: {e}")
                count += 1
                count = await restart_browser_profile(driver, env_id, unique_id, env_name, count)
                continue

    except Exception as e:
        logger.error(
            f" (main_flow) ERROR in Profile №: {unique_id}, Env_Name: {env_name}, Env ID: {env_id}, Ошибка: {e}"
        )
        raise MainError(f"Main flow failed: {e}")

    finally:
        # Завершаем профиль корректно
        if driver and mode_close_profile_or_not.lower() == "y":
            try:
                close_tabs = MetaMaskHelper(driver)
                close_tabs.delete_others_windows()
                driver.quit()
                await BrowserManager.stop_browser_profile(env_id)
                # Добавляем сообщение о завершении работы с профилем
                logger.warning(
                    f"\n{'#' * 20} (main_flow) SCRIPT ENDED for Profile №: {unique_id}, Env_Name: {env_name}, Env ID: {env_id} {'#' * 20}\n"
                )
            except Exception as e:
                logger.error(f"Error closing browser profile: {e}")


async def read_user_list_file(
    worksheet_mm, start_account, end_account, mix_profiles, workbook_mm
):
    """Чтение данных из Excel-файла."""
    profiles = []
    logger.debug(
        f"[DATABASE READ] Чтение из базы данных Excel-файла. От {start_account} до {end_account}."
    )

    try:
        for row_start_with_1nd_line in range(start_account, end_account + 1):
            row = row_start_with_1nd_line + 1  # Прибавляем 1, так как в Excel профиля начинаются с 2 строки
            try:
                # Чтение данных из базы
                unique_id = worksheet_mm.cell(row=row, column=1).value
                password = worksheet_mm.cell(row=row, column=2).value
                seed = worksheet_mm.cell(row=row, column=3).value
                mm_address = worksheet_mm.cell(row=row, column=4).value
                private_key = worksheet_mm.cell(row=row, column=5).value

                # Проверка и создание пароля если отсутствует
                if password is None and seed:
                    password = create_password()
                    worksheet_mm.cell(row=row, column=2).value = password
                    workbook_mm.save(DATA_BASE_PATH)
                    logger.info(
                        f"[NEW PASSWORD] Для Профиль № {unique_id} создан новый пароль."
                    )

                # Проверка и создание нового кошелька если отсутствует seed
                if seed is None:
                    seed, mm_address, private_key = create_wallet()
                    password = create_password()  # Создаем новый пароль для нового кошелька
                    worksheet_mm.cell(row=row, column=2).value = password
                    worksheet_mm.cell(row=row, column=3).value = seed
                    worksheet_mm.cell(row=row, column=4).value = mm_address
                    worksheet_mm.cell(row=row, column=5).value = private_key
                    workbook_mm.save(DATA_BASE_PATH)
                    logger.info(
                        f"[NEW WALLET] Для Профиль № {unique_id} создан новый кошелек:\n"
                        f"Адрес: {mm_address}\n"
                        f"Seed: {seed}\n"
                        f"Private Key: {private_key}\n"
                        f"Пароль: {password}"
                    )

                # Проверка наличия адреса кошелька
                if mm_address is None:
                    logger.warning(
                        f"[MISSING ADDRESS] В базе отсутствует адрес ETH wallet для Профиль № {unique_id}.\n"
                        f"При успешном входе в профиль и MetaMask, адрес будет автоматически добавлен в базу."
                    )

                # Добавление профиля в список
                profiles.append(
                    [
                        unique_id,
                        password,
                        seed,
                        mm_address,
                        private_key,
                        worksheet_mm,
                        workbook_mm,
                        row_start_with_1nd_line,
                    ]
                )
                logger.debug(f"[PROFILE ADDED] Профиль № {unique_id} добавлен в список для обработки")

            except Exception as e:
                logger.error(f"[ERROR] Ошибка чтения данных для строки {row}: {traceback.format_exc()}")
                continue

        if mix_profiles.lower() == "y":
            random.shuffle(profiles)
            logger.info(f"[SHUFFLE] Профили перемешаны.")

        if not profiles:
            raise MainError("No valid profiles found in the specified range")
        return profiles

    except Exception as e:
        logger.error(f"[CRITICAL ERROR] Ошибка чтения списка пользователей: {e}")
        raise MainError(f"Failed to read user list: {e}")



def get_user_input():
    """Получение параметров от пользователя с улучшенной обработкой ввода и логированием."""
    try:
        start_account = int(input("\nВведите номер начального профиля: "))
        end_account = int(input("Введите номер конечного профиля: "))

        if start_account < 1 or end_account < start_account:
            raise ValueError("Invalid account range")

        # Получаем значение из конфига
        default_close_profile = MODE_CLOSE_PROFILE
        logger.debug(f"Используется значение MODE_CLOSE_PROFILE из config.yaml: {default_close_profile}")

        default_prompt = "Y" if default_close_profile else "N"

        mode_close_profile_or_not = input(
            f"\nЗакрывать профиль после выполнения? Y/N (по умолчанию: {default_prompt}): "
        ).strip().lower()

        # Если пользователь не ввел значение, используем значение из конфига
        if not mode_close_profile_or_not:
            mode_close_profile_or_not = "y" if default_close_profile else "n"
            logger.info(f"Используется значение из конфига: {mode_close_profile_or_not.upper()}")
        else:
            logger.info(f"Используется значение, введенное пользователем: {mode_close_profile_or_not.upper()}")

        delay_from_to = None
        mix_profiles = "n"

        if start_account != end_account:
            # Получаем значение из конфига для перемешивания профилей
            default_mix_profiles = MIX_PROFILES
            default_mix_prompt = "Y" if default_mix_profiles else "N"

            mix_profiles = input(
                f"\nПеремешивать кошельки? Y/N (по умолчанию: {default_mix_prompt}): "
            ).strip().lower()

            # Если пользователь не ввел значение, используем значение из конфига
            if not mix_profiles:
                mix_profiles = "y" if default_mix_profiles else "n"
                logger.info(f"Используется значение из конфига для перемешивания: {mix_profiles.upper()}")
            else:
                logger.info(f"Используется значение, введенное пользователем для перемешивания: {mix_profiles.upper()}")

            if mix_profiles == "y":
                print("Профили будут перемешаны перед запуском.")
            else:
                print("Профили будут запущены в порядке их создания.")

            # Получаем настройки задержки из конфига

            default_delay_enabled = PROFILE_DELAY.get('ENABLED', False)
            default_delay_min = PROFILE_DELAY.get('MIN', 10)
            default_delay_max = PROFILE_DELAY.get('MAX', 60)

            delay_prompt = "Y" if default_delay_enabled else "N"
            delay_execution = input(
                f"\nДобавить задержку между профилями? Y/N (по умолчанию: {delay_prompt})\n"
            ).strip().lower()

            if not delay_execution:
                delay_execution = "y" if default_delay_enabled else "n"
                logger.info(f"Используется значение из конфига для задержки: {delay_execution.upper()}")

            if delay_execution == "y":
                while True:
                    try:
                        delay_input = input(
                            f"Введите задержку выполнения (от, до) секунд, через запятую (по умолчанию: {default_delay_min},{default_delay_max}): "
                        ).strip()

                        if not delay_input:
                            delay_from_to = [str(default_delay_min), str(default_delay_max)]
                            logger.info(f"Используется значение задержки из конфига: {default_delay_min}-{default_delay_max} секунд")
                            break

                        delay_from_to = delay_input.split(",")
                        if len(delay_from_to) == 2 and float(delay_from_to[0]) < float(delay_from_to[1]):
                            print(f"Задержка установлена в диапазоне от {delay_from_to[0]} до {delay_from_to[1]} секунд.")
                            break
                        else:
                            logger.error("Не корректный диапазон. Повторите ввод, например: 10,60")
                            time.sleep(0.2)
                    except ValueError:
                        logger.error("Ошибка! Введите корректные числовые значения")
                        time.sleep(0.2)
            else:
                print("Задержка отсутствует.")

        return (
            start_account,
            end_account,
            mix_profiles,
            delay_from_to,
            mode_close_profile_or_not,
        )
    except ValueError as e:
        logger.error(f"Invalid input: {e}")
        raise MainError(f"Invalid user input: {e}")
    except Exception as e:
        logger.error(f"Error getting user input: {e}")
        logger.error(traceback.format_exc())  # Получаем полный traceback
        raise MainError(f"Failed to get user input: {e}")


async def operationEnv(
    driver, seed, env_id, password, mm_address, worksheet_mm, workbook_mm, row, file_path
):
    """Основная операция."""
    try:
        logger.debug(
            f" (operationEnv) STEP 2 <<< Запуск Env ID: {env_id} Основная операция >>>"
        )
        driver.refresh()
        driver.maximize_window()
        await asyncio.sleep(1)

        # Используем методы через экземпляр MetaMaskHelper
        helper = MetaMaskHelper(driver)
        helper.delete_others_windows()

        if modify_file_runtimelavamoat(env_id):
            try:
                wallet_mm_from_browser_extension = helper.meta_mask(
                    seed,
                    password,
                    mm_address,
                    row,
                    workbook_mm,
                    worksheet_mm,
                    file_path,
                )
                logger.debug(
                    f"wallet_mm_from_browser_extension: {wallet_mm_from_browser_extension}, type: {type(wallet_mm_from_browser_extension)}"
                )

                # Проверка БД на предмет наступления времени в необходимости выполнения активности faucet_morkie
                # и занесением результата в БД.
                try:
                    if wallet_mm_from_browser_extension:
                        process_activity(driver, wallet_mm_from_browser_extension, row)
                    else:
                        process_activity(driver, mm_address, row)
                except DatabaseError as e:
                    logger.error(f"Database error in operationEnv: {e}")
                    raise
                except Exception as e:
                    logger.error(f"Error processing activity: {e}")
                    raise

                # Открываем вкладки для проверки активов по адресу кошелька в Debank и MonadExplorer.
                helper.open_tab(f"https://testnet.monadexplorer.com/address/{wallet_mm_from_browser_extension}")
                helper.open_tab("https://debank.com/profile/" + wallet_mm_from_browser_extension)

            except Exception as e:
                logger.error(f"Error in MetaMask operation: {e}")
                raise

    except Exception as e:
        logger.error(f"Error in operationEnv: {e}")
        raise MainError(f"Operation failed: {e}")


async def main():
    """Главная функция"""
    script_start = datetime.now()
    logger.info("Начало работы скрипта")

    # Инициализируем переменные в начале функции
    count_profile = 0
    profiles = []

    try:
        # Проверяем режим работы
        if GLOBAL_SETTINGS.get('AUTO_MODE', False):  # Если включен автоматический режим в config.yaml

            # logger.info("Установка в расписание для следующего запуска в автоматическом режиме.")
            schedule_next_run()  # Установка в расписание для следующего запуска в автоматическом режиме.

            # Получаем один рандомный аккаунт
            logger.info("Автоматический режим: выбор аккаунта из базы данных")
            selected_account, wallet = process_random_profile()
            start_account = selected_account
            end_account = selected_account
            mode_close_profile_or_not = 'y'
            mix_profiles = 'n'
            delay_from_to = [0,0]
        else:
            # Интерактивный режим. Получаем параметры от пользователя
            start_account, end_account, mix_profiles, delay_from_to, mode_close_profile_or_not = get_user_input()

        # Проверка существует ли файл с базой данных DB.xlsx
        if os.path.exists(DATA_BASE_PATH):
            logger.debug(f" (main) DATABASE exists: {DATA_BASE_PATH}.")
            workbook_mm, worksheet_mm = workbook_worksheet()
        else:
            logger.info(f" (main) DATABASE База данных отсутствует. Создание нового файла базы данных.")
            workbook_mm = Workbook()
            worksheet_mm = workbook_mm.active
            worksheet_mm.title = WORKSHEET_NAME
            worksheet_mm["A1"] = "No"
            worksheet_mm["B1"] = "Password"
            worksheet_mm["C1"] = "Mnemonic"
            worksheet_mm["D1"] = "Address"
            worksheet_mm["E1"] = "Private key"

            # Запись числовых значений от 1 до 100 в первую колонку, начиная со 2-й строки.
            for row in range(2, 102):  # от 2 до 101 (включительно)
                worksheet_mm.cell(row=row, column=1)

            workbook_mm.save(DATA_BASE_PATH)  # Сохранение нового файла
            logger.info(f" (main) DATABASE created new file: {DATA_BASE_PATH}.")
            # код для работы с worksheet_mm загружаем его
            workbook_mm, worksheet_mm = workbook_worksheet()

        # Получаем список профилей
        profiles = await read_user_list_file(
            worksheet_mm, start_account, end_account, mix_profiles, workbook_mm
        )

        if not profiles:
            raise MainError("No valid profiles found in the file")

        # Перемешиваем профили если нужно
        if mix_profiles == "y":
            random.shuffle(profiles)
            logger.info("Profiles have been shuffled")

        logger.info(f"Будет обработано профилей: {len(profiles)}")

        for idx, profile in enumerate(profiles, 1):
            try:
                (
                    unique_id, password, seed, mm_address,
                    private_key, worksheet_mm, workbook_mm, row
                ) = profile

                env_id, unique_id, env_name = await BrowserManager.get_list_browser_profiles(
                    unique_id
                )

                start_time = datetime.now()
                logger.update(
                    f"\n{'=' * 80}\n"
                    f"Обработка Профиль № {unique_id} ({idx}/{len(profiles)})\n"
                    f"Имя: {env_name}, Адрес: {mm_address}\n"
                    f"{'=' * 80}\n"
                )

                await main_flow(
                    env_id,
                    seed,
                    password,
                    env_name,
                    unique_id,
                    mm_address,
                    worksheet_mm,
                    workbook_mm,
                    row,
                    mode_close_profile_or_not
                )

                duration = datetime.now() - start_time
                logger.info(f"Профиль № {unique_id} обработан за {duration}")
                count_profile += 1

                # Задержка между профилями
                if delay_from_to and idx < len(profiles):
                    delay = random.uniform(float(delay_from_to[0]), float(delay_from_to[1]))
                    logger.info(f"Пауза {delay:.1f} сек...")
                    time.sleep(delay)

            except Exception as e:
                # Используем profile[0] как запасной вариант для unique_id
                profile_id = profile[0] if profile else f"profile_{idx}"
                logger.error(f"Error processing Профиль № {profile_id}: {e}")
                continue

    except MainError as e:
        logger.error(f"Critical error in main: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Unexpected error in main: {e}")
        sys.exit(1)
    finally:
        total_time = datetime.now() - script_start
        logger.info(f"\nСкрипт завершен. Общее время: {total_time}")

        # Формируем правильное окончание для слова "профиль"
        if count_profile % 10 == 1 and count_profile % 100 != 11:
            profile_word = "профиль"
        elif 2 <= count_profile % 10 <= 4 and (count_profile % 100 < 10 or count_profile % 100 >= 20):
            profile_word = "профиля"
        else:
            profile_word = "профилей"

        logger.info(f"\n Успешно обработано {count_profile} {profile_word} из {len(profiles)}\n")


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("\nСкрипт остановлен пользователем")
        sys.exit(0)
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        sys.exit(1)
